import openpyxl as op
import datetime as dt

""" 
CRUD Basico - Excel (BD)
"""

# Funcion que crea un archivo xlsx y agrega encabezado desde una lista
def crear_info(archivo, datos):
    libro = op.Workbook()
    hoja = libro.active
    col = 1
    for i in datos:
        hoja.cell(row=1, column=col, value=i)
        col += 1
    libro.save(archivo)
    return f"{archivo} fue creada exitosamente"

# Funcion que agrega datos a la BD (archivo xlsx)
def create(archivo, datos):
    libro = op.load_workbook(archivo)
    hoja = libro.active
    fila = hoja.max_row + 1
    col = 1
    fecha = dt.datetime.now()
    for i in datos:
        hoja.cell(row=fila, column=col, value=i)
        col += 1
    hoja.cell(row=fila, column=hoja.max_column, value=fecha)
    libro.save(archivo)
    return f"Registro agregado correctamente a {archivo}"

# Funcion que lee datos en una BD (archivo xlsx)
def read(archivo, info):
    libro = op.load_workbook(archivo)
    hoja = libro.active
    datos = dict()
    enc = list()
    for row in hoja.iter_rows(min_row=1, max_col=hoja.max_column, max_row=1):
            for celda in row:
                enc.append(celda.value) 
    if info.lower() == "todos":          
        for row in hoja.iter_rows(min_row=2, max_col=hoja.max_column, max_row=hoja.max_row):
            info = dict()
            for i in range(0, len(row)):
                info[enc[i]] = row[i].value
            datos[row[0].value] = info
    else:
        for row in hoja.iter_rows(min_row=2, max_col=hoja.max_column, max_row=hoja.max_row):
            if row[0].value == info:
                info = dict()
                for i in range(0, len(row)):
                    info[enc[i]] = row[i].value
                datos[row[0].value] = info
    return datos

def update(archivo, datos):
    libro = op.load_workbook(archivo)
    hoja = libro.active
    fecha = dt.datetime.now()
    for row in hoja.iter_rows(min_row=2, max_col=hoja.max_column, max_row=hoja.max_row):
            if row[0].value == datos[0]:               
                for i in range(0, len(datos)):
                    row[i].value = datos[i]               
                row[hoja.max_column - 1].value = fecha                            
    libro.save(archivo)

def delete(archivo, id):
    libro = op.load_workbook(archivo)
    hoja = libro.active
    for row in hoja.iter_rows(min_row=2, max_col=hoja.max_column, max_row=hoja.max_row):
            if row[0].value == id:
                for celda in row:
                    celda.value = ""
    libro.save(archivo)
                
               

#crear_info("archivo.xlsx", ["id", "nombres", "apellidos", "telefono", "Fecha_modificacion"])
#create("archivo.xlsx", ["123445", "Joel", "Guerrero", "3178500744"])
#create("archivo.xlsx", ["234556", "Sandra", "Romero", "3174354213"])
#print(read("archivo.xlsx", info="Todos"))
#update("archivo.xlsx", ["123445", "Jose Luis", "Guerrero", "3178500744"])
delete("archivo.xlsx", id="234556")