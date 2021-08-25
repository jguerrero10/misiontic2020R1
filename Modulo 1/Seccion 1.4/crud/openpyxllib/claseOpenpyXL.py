import openpyxl

wb = openpyxl.Workbook() # Crea un libro de trabajo

hoja = wb.active # Crea hoja por defecto
ws1 = wb.create_sheet("hoja_nueva") # Crea hoja nueva
ws2 = wb.create_sheet("hoja_principla", 0) # Crea hoja nueva en la primera posicion
ws2 = wb.create_sheet("hoja_principla", -1) # Crea hoja nueva en la primera posicion

hoja.title = "Titulo Nuevo" # Cambiar el nombre del libro

hoja.sheet_properties.tabColor = "1072BA" #Color de fondo de la pestaña

hoja = wb['Titulo Nuevo'] # Obtenerla con clave del libro de trabajo

print(wb.sheetnames) # Obtener los nombres de las hojas del libro

for h in wb: # Recorre las hojas de trabajo
    print(h.title)

# Realizar una copia de hojas de trabajo dentro de un solo libro
source = wb.active 
target = wb.copy_worksheet(source)

c = hoja['A4'] # Acceder directamente  a la celda

hoja['A4'] = "Hola"  # Asignar valor directamente a una celda
hoja['A5'] = "mundo"

d = hoja.cell(row=4, column=2, value=10) #Se pueda asigna valor tambien asi

cell_range = hoja['A1':'C2'] # rango de celdas
colC = hoja['C:D'] # Rango de Columnas
row = hoja[10] #Rango de filas

for row in hoja.iter_rows(min_row=1, max_col=3, max_row=2): # Devuelve las filas
    for cell in row:
        print(cell)

for col in hoja.iter_cols(min_row=1, max_col=3, max_row=2): # Devuelve las columnas
    for cell in col:
        print(cell)

print(tuple(hoja.rows)) # Recorre todas las filas 
print(tuple(hoja.columns)) # Recorre todas las columnas

for row in hoja.values:
    for value in row:
        print(value)

wb.save('archivo.xlsx') # Guardar un archivo
# wb2 = openpyxl.load_workbook('archivo.xlsx')

